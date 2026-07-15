# ==========================================================
# APP DAE - CONVERSOR DE AVE
# Versión 1.5
# ==========================================================
import streamlit as st
import pandas as pd
import numpy as np
import unicodedata
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import (
  PatternFill,
  Border,
  Side,
  Font,
  Alignment
)
from openpyxl.utils.dataframe import dataframe_to_rows
# ==========================================================
# CONFIGURACIÓN STREAMLIT
# ==========================================================
st.set_page_config(
  page_title="DAE - Conversor",
  page_icon="📄",
  layout="wide"
)
# ==========================================================
# LOGO
# ==========================================================
st.image(
  "images.png",
  width=100
)
st.title("Conversor de AVE")
st.caption(
  "Procesamiento automático de participantes, planta y resultados."
)
st.divider()
# ==========================================================
# SESSION STATE
# ==========================================================
VARIABLES = [
  "reporte",
  "archivo",
  "resumen",
  "inconsistencias"
]
for v in VARIABLES:
  if v not in st.session_state:
      st.session_state[v] = None
# ==========================================================
# COLORES EXCEL
# ==========================================================
COLOR_AMARILLO = "FFF59D"
COLOR_ROJO = "EF9A9A"
COLOR_VERDE = "C8E6C9"
COLOR_GRIS = "E0E0E0"
# ==========================================================
# NORMALIZAR TEXTO
# ==========================================================
def normalizar_texto(valor):
  if pd.isna(valor):
      return ""
  valor = str(valor)
  valor = valor.upper()
  valor = valor.strip()
  valor = unicodedata.normalize(
      "NFKD",
      valor
  )
  valor = "".join(
      c for c in valor
      if not unicodedata.combining(c)
  )
  valor = re.sub(
      r"\s+",
      " ",
      valor
  )
  return valor
# ==========================================================
# NORMALIZAR CÉDULA
# ==========================================================
def normalizar_id(valor):
  if pd.isna(valor):
      return ""
  valor = str(valor)
  valor = valor.replace(".0", "")
  valor = valor.strip()
  if valor.lower() == "nan":
      return ""
  return valor
# ==========================================================
# BUSCAR COLUMNA AUTOMÁTICAMENTE
# ==========================================================
def buscar_columna(df, opciones):
  columnas = {
      normalizar_texto(c): c
      for c in df.columns
  }
  for opcion in opciones:
      opcion = normalizar_texto(opcion)
      if opcion in columnas:
          return columnas[opcion]
  return None
# ==========================================================
# OBTENER COLUMNA OBLIGATORIA
# ==========================================================
def obtener_columna(df, opciones):
  columna = buscar_columna(
      df,
      opciones
  )
  if columna is None:
      raise Exception(
          f"No se encontró ninguna columna entre: {opciones}"
      )
  return columna
# ==========================================================
# CARGAR EXCEL
# ==========================================================
def cargar_excel(archivo):
  try:
      return pd.read_excel(
          archivo,
          dtype=object
      )
  except Exception as e:
      st.error(e)
      st.stop()
# ==========================================================
# CREAR OBSERVACIONES
# ==========================================================
def crear_observaciones(df):
  if "Observaciones" not in df.columns:
      df["Observaciones"] = ""
  return df
# ==========================================================
# AGREGAR OBSERVACIÓN
# ==========================================================
def agregar_observacion(df, mascara, texto):
  existente = df.loc[mascara, "Observaciones"].fillna("")
  ya_termina_en_separador = existente.str.rstrip().str.endswith(";")
  necesita_separador = (
      (existente.str.len() > 0) & (~ya_termina_en_separador)
  )
  con_separador = existente.where(
      ~necesita_separador,
      existente + "; "
  )
  df.loc[mascara, "Observaciones"] = con_separador + texto + "; "
  return df
# ==========================================================
# PREPARAR PARTICIPANTES
# ==========================================================
def preparar_participantes(df):
  df = df.copy()
  col_nombre = obtener_columna(
      df,
      [
          "Nombre"
      ]
  )
  col_apellidos = obtener_columna(
      df,
      [
          "Apellido(s)",
          "Apellidos",
          "Apellido"
      ]
  )
  col_id = obtener_columna(
      df,
      [
          "Número de ID",
          "Numero de ID",
          "Cedula",
          "Cédula",
          "Documento",
          "Documento de identidad"
      ]
  )
  col_cargo = buscar_columna(
      df,
      [
          "Departamento",
          "Cargo"
      ]
  )
  col_institucion = buscar_columna(
      df,
      [
          "Institución",
          "Institucion"
      ]
  )
  col_correo = buscar_columna(
      df,
      [
          "Correo",
          "Email",
          "Correo electrónico",
          "E-mail",
          "Dirección Email"
      ]
  )
  participantes = pd.DataFrame()
  participantes["Nombre"] = (
      df[col_nombre]
      .fillna("").astype(str)
      + " "
      + df[col_apellidos].fillna("").astype(str)
      ).apply(normalizar_texto)
  participantes["NumeroID"] = (
      df[col_id]
      .apply(normalizar_id)
  )
  if col_cargo:
      participantes["Cargo"] =  (
          df[col_cargo]
          .apply(normalizar_texto)
      )
  else:
      participantes["Cargo"] = ""
  participantes["Dependencia"] = ""
  participantes["Seccional"] = ""
  participantes["Institución"] = (
      df[col_institucion].apply(normalizar_texto)
      if col_institucion else ""
  )
  participantes["Correo"] = (
      df[col_correo].apply(normalizar_texto)
      if col_correo else ""
  )
  participantes = crear_observaciones(participantes)
  return participantes
# ==========================================================
# PREPARAR PLANTA
# ==========================================================
def preparar_planta(df):
   df = df.copy()
   col_nombre = obtener_columna(df, ["Nombre"])
   col_id = obtener_columna(df, ["Cedula", "Cédula"])
   col_dependencia = obtener_columna(
       df,
       ["NombreDependencia", "Dependencia"]
   )
   col_seccional = obtener_columna(
       df,
       ["Seccional Nor", "Seccional"]
   )
   planta = pd.DataFrame()
   planta["Nombre"] = df[col_nombre].apply(normalizar_texto)
   planta["NumeroID"] = df[col_id].apply(normalizar_id)
   planta["Dependencia"] = df[col_dependencia]
   planta["Seccional"] = df[col_seccional]
   planta = planta.drop_duplicates()
   return planta
# ==========================================================
# PREPARAR CALIFICACIONES
# ==========================================================
def preparar_calificaciones(df):
  df = df.copy()
  col_nombre = obtener_columna(
      df,
      [
          "Nombre Completo",
          "Nombre"
      ]
  )
  col_id = buscar_columna(
      df,
      [
          "Número de ID",
          "Numero de ID",
          "Cedula",
          "Cédula"
      ]
  )
  col_nota = obtener_columna(
      df,
      [
          "Total del curso (Real)",
          "Nota",
          "Calificación",
          "Calificacion"
      ]
  )
  resultados = pd.DataFrame()
  resultados["Nombre"] = (
      df[col_nombre]
      .apply(normalizar_texto)
  )
  notas_crudas = df[col_nota].apply(
      lambda v: 0 if str(v).strip() == "-" else v
  )
  if col_id:
      resultados["NumeroID"] = (
          df[col_id]
          .apply(normalizar_id)
      )
  else:
      resultados["NumeroID"] = ""
  resultados["Nota"] = pd.to_numeric(
      notas_crudas,
      errors="coerce"
  )
  col_correo = buscar_columna(
      df,
      [
          "Correo",
          "Email",
          "Correo electrónico",
          "E-mail",
            "Dirección Email"
      ]
  )
  resultados["Correo"] = (
      df[col_correo].apply(normalizar_texto)
      if col_correo else ""
  )
  return resultados
# ==========================================================
# PREPARAR APROBADOS
# ==========================================================
def preparar_aprobados(df):
  df = df.copy()
  col_nombre = obtener_columna(
      df,
      [
          "Nombre Completo",
          "Nombre"
      ]
  )
  col_id = buscar_columna(
      df,
      [
          "Número de ID",
          "Numero de ID",
          "Cedula",
          "Cédula"
      ]
  )
  aprobados = pd.DataFrame()
  aprobados["Nombre"] = (
      df[col_nombre]
      .apply(normalizar_texto)
  )
  if col_id:
      aprobados["NumeroID"] = (
          df[col_id]
          .apply(normalizar_id)
      )
  else:
      aprobados["NumeroID"] = ""
  col_correo = buscar_columna(
      df,
      [
          "Correo",
          "Email",
          "Correo electrónico",
          "E-mail",
            "Dirección Email"
      ]
  )
  aprobados["Correo"] = (
      df[col_correo].apply(normalizar_texto)
      if col_correo else ""
  )
  return aprobados
# ==========================================================
# CRUCE CON PLANTA
# ==========================================================
def enriquecer_con_planta(participantes, planta):
   """
   Cruza participantes con Planta para completar Dependencia/Seccional
   y recuperar cédulas faltantes.
   1) Cruce por NumeroID (dict O(1))
   2) Cruce por Nombre, solo para pendientes y solo si el nombre
      es único en Planta (evita asignar mal por homónimos)
   3) Marca no encontrados y nombres duplicados en Planta
   """
   df = participantes.copy()
   estadisticas = {
       "Encontrados por ID": 0,
       "Encontrados por Nombre": 0,
       "Cedulas Recuperadas": 0,
       "No encontrados en Planta": 0,
       "Nombres Duplicados en Planta": 0

   }

   # ======================================================
   # DICCIONARIOS DE BÚSQUEDA
   # ======================================================
   planta_id = planta.drop_duplicates(subset="NumeroID")
   dict_por_id = planta_id.set_index("NumeroID")[
       ["Dependencia", "Seccional"]
   ].to_dict(orient="index")
   conteo_nombres = planta.groupby("Nombre").size()
   nombres_unicos = set(conteo_nombres[conteo_nombres == 1].index)
   nombres_duplicados = set(conteo_nombres[conteo_nombres > 1].index)
   planta_nombre = planta[planta["Nombre"].isin(nombres_unicos)]
   dict_por_nombre = planta_nombre.set_index("Nombre")[
       ["NumeroID", "Dependencia", "Seccional"]
   ].to_dict(orient="index")
   # ======================================================
   # 1. CRUCE POR CÉDULA
   # ======================================================
   encontrados_id = df["NumeroID"].apply(
       lambda nid: dict_por_id.get(nid) if nid != "" else None
   )
   mascara_id = encontrados_id.notna()
   df.loc[mascara_id, "Dependencia"] = encontrados_id[mascara_id].apply(
       lambda x: x["Dependencia"]
   )
   df.loc[mascara_id, "Seccional"] = encontrados_id[mascara_id].apply(
       lambda x: x["Seccional"]
   )
   estadisticas["Encontrados por ID"] = int(mascara_id.sum())
   agregar_observacion(df, mascara_id, "Encontrado en Planta por ID")
   # ======================================================
   # 2. CRUCE POR NOMBRE (solo pendientes)
   # ======================================================
   pendientes = ~mascara_id
   encontrados_nombre = pd.Series(None, index=df.index, dtype=object)
   encontrados_nombre[pendientes] = df.loc[pendientes, "Nombre"].map(
       dict_por_nombre
   )
   mascara_nombre = encontrados_nombre.notna()
   estadisticas["Encontrados por Nombre"] = int(mascara_nombre.sum())
   df.loc[mascara_nombre, "Dependencia"] = encontrados_nombre[
       mascara_nombre
   ].apply(lambda x: x["Dependencia"])
   df.loc[mascara_nombre, "Seccional"] = encontrados_nombre[
       mascara_nombre
   ].apply(lambda x: x["Seccional"])
   agregar_observacion(df, mascara_nombre, "Encontrado en Planta por Nombre")
   # Recuperar cédula solo si venía vacía
   mascara_recuperar = mascara_nombre & (df["NumeroID"] == "")
   df.loc[mascara_recuperar, "NumeroID"] = encontrados_nombre[
       mascara_recuperar
   ].apply(lambda x: x["NumeroID"])
   estadisticas["Cedulas Recuperadas"] = int(mascara_recuperar.sum())
   agregar_observacion(df, mascara_recuperar, "Cedula recuperada desde Planta")
   # ======================================================
   # 3. NO ENCONTRADOS EN PLANTA
   # ======================================================
   sin_dependencia = df["Dependencia"].fillna("").eq("")
   estadisticas["No encontrados en Planta"] = int(sin_dependencia.sum())
   agregar_observacion(df, sin_dependencia, "No encontrado en Planta")
   # ======================================================
   # 4. NOMBRES DUPLICADOS EN PLANTA (entre los no encontrados)
   # ======================================================
   mascara_dup = df["Nombre"].isin(nombres_duplicados) & sin_dependencia
   estadisticas["Nombres Duplicados en Planta"] = int(mascara_dup.sum())
   agregar_observacion(df, mascara_dup, "Nombre duplicado en Planta")
   # ======================================================
   # 5. RESPALDO: USAR INSTITUCIÓN SI NO SE ENCONTRÓ EN PLANTA
   # ======================================================
   if "Institución" in df.columns:
       usar_institucion = sin_dependencia & (df["Institución"] != "")
       df.loc[usar_institucion, "Dependencia"] = df.loc[
           usar_institucion, "Institución"
       ]
   return df, estadisticas


# ==========================================================
# VALIDAR IDs VACÍOS
# ==========================================================
def validar_ids_vacios(df):
  df = df.copy()
  mascara = df["NumeroID"] == ""
  agregar_observacion(
      df,
      mascara,
      "ID vacío"
  )
  return df, int(mascara.sum())
# ==========================================================
# VALIDAR DUPLICADOS
# ==========================================================
def validar_duplicados(df):
  df = df.copy()
  mascara = (
      df["NumeroID"] != ""
  ) & (
      df["NumeroID"].duplicated(keep=False)
  )
  agregar_observacion(
      df,
      mascara,
      "Cédula duplicada"
  )
  duplicados = df.loc[
      mascara,
      "NumeroID"
  ].unique()
  return df, list(duplicados)
# ==========================================================
# APROBADOS DUPLICADOS
# ==========================================================
def contar_aprobados_duplicados(df):
  mascara = (
      df["Aprobó"] == "Sí"
  ) & (
      df["NumeroID"].duplicated(
          keep=False
      )
  )
  return int(
      df.loc[
          mascara,
          "NumeroID"
      ].nunique()
  )
# ==========================================================
# LIMPIAR OBSERVACIONES
# ==========================================================
def limpiar_observaciones(df):
  df = df.copy()
  df["Observaciones"] = (
      df["Observaciones"]
      .str.replace(
          ";;",
          ";",
          regex=False
      )
      .str.strip()
      .str.rstrip(";")
  )
  return df
# ==========================================================
# VALIDACIÓN GENERAL
# ==========================================================
def ejecutar_validaciones(df):
  df, ids_vacios = validar_ids_vacios(df)
  df, duplicados = validar_duplicados(df)
  df = limpiar_observaciones(df)
  estadisticas = {
      "IDs Vacíos": ids_vacios,
      "Duplicados": len(duplicados),
      "Lista Duplicados": duplicados
  }
  return df, estadisticas
# ==========================================================
# CRUCE CON CALIFICACIONES
# ==========================================================
def procesar_con_nota(participantes, resultados):
  df = participantes.copy()
  if "Correo" not in df.columns:
      df["Correo"] = ""
  resultados = resultados.copy()
  if "Correo" not in resultados.columns:
      resultados["Correo"] = ""
  #--------------------------------------------------------
  # DETECTAR CÉDULAS CON NOTAS EN CONFLICTO
  # (mismo NumeroID, más de una Nota distinta -> personas
  # distintas con la misma cédula mal cargada, o duplicados reales)
  #--------------------------------------------------------
  validos = resultados[resultados["NumeroID"] != ""]
  conflictivas = set(
      validos.groupby("NumeroID")["Nota"]
      .nunique(dropna=False)
      .loc[lambda s: s > 1]
      .index
  )
  #--------------------------------------------------------
  # DICCIONARIO PRECISO (NumeroID, Correo) -> Nota
  # Se usa para resolver los casos en conflicto
  #--------------------------------------------------------
  dict_id_correo = (
      resultados.drop_duplicates(subset=["NumeroID", "Correo"])
      .set_index(["NumeroID", "Correo"])["Nota"]
      .to_dict()
  )
  #--------------------------------------------------------
  # DICCIONARIO POR NumeroID, solo para IDs SIN conflicto
  #--------------------------------------------------------
  sin_conflicto = resultados[~resultados["NumeroID"].isin(conflictivas)]
  dict_id = (
      sin_conflicto.drop_duplicates(subset="NumeroID")
      .set_index("NumeroID")["Nota"]
      .to_dict()
  )
  #--------------------------------------------------------
  # DICCIONARIO POR Nombre (último recurso, sin cédula)
  #--------------------------------------------------------
  dict_nombre = (
      resultados.drop_duplicates(subset="Nombre")
      .set_index("Nombre")["Nota"]
      .to_dict()
  )
  #--------------------------------------------------------
  # ASIGNAR NOTA SEGÚN PRIORIDAD:
  # 1) ID conflictivo -> exigir match exacto (ID, Correo)
  # 2) ID sin conflicto -> buscar por ID
  # 3) sin match por ID -> buscar por Nombre
  #--------------------------------------------------------
  es_conflictivo = df["NumeroID"].isin(conflictivas)
  nota_por_id_correo = pd.Series(
      list(zip(df["NumeroID"], df["Correo"])),
      index=df.index
  ).map(dict_id_correo)
  nota_por_id = df["NumeroID"].map(dict_id)
  nota_por_nombre = df["Nombre"].map(dict_nombre)
  nota_final = nota_por_id.where(nota_por_id.notna(), nota_por_nombre)
  nota_final = nota_final.mask(es_conflictivo, nota_por_id_correo)
  df["Nota"] = pd.to_numeric(nota_final, errors="coerce")
  #--------------------------------------------------------
  # LO QUE SIGUE SIN RESOLVERSE (conflicto + correo no coincide)
  # se marca para revisión manual, nunca se adivina
  #--------------------------------------------------------
  mascara_inconsistente = es_conflictivo & nota_por_id_correo.isna()
  #--------------------------------------------------------
  # APROBADOS
  #--------------------------------------------------------
  df["Aprobó"] = np.where(
      df["Nota"] >= 3.5,
      "Sí",
      "No"
  )
  df.loc[mascara_inconsistente, "Aprobó"] = "Revisar"
  df.loc[mascara_inconsistente, "Nota"] = np.nan
  agregar_observacion(
      df,
      mascara_inconsistente,
      "Nota inconsistente entre registros - revisar manualmente"
  )
  return df
# ==========================================================
# CURSOS SIN NOTA
# ==========================================================
def procesar_sin_nota(participantes, aprobados):
 df = participantes.copy()
 if "Correo" not in df.columns:
     df["Correo"] = ""
 aprobados = aprobados.copy()
 if "Correo" not in aprobados.columns:
     aprobados["Correo"] = ""
 #--------------------------------------------------------
 # CÉDULAS DUPLICADAS EN PARTICIPANTES
 # (misma cédula para más de una persona -> ambiguo)
 #--------------------------------------------------------
 conteo_participantes = df.loc[df["NumeroID"] != "", "NumeroID"].value_counts()
 cedulas_duplicadas = set(conteo_participantes[conteo_participantes > 1].index)
 es_conflictivo = df["NumeroID"].isin(cedulas_duplicadas)
 #--------------------------------------------------------
 # MATCH EXACTO (NumeroID, Correo) -> para resolver conflictos
 #--------------------------------------------------------
 set_id_correo = set(zip(aprobados["NumeroID"], aprobados["Correo"]))
 match_id_correo = pd.Series(
     list(zip(df["NumeroID"], df["Correo"])),
     index=df.index
 ).isin(set_id_correo)
 #--------------------------------------------------------
 # MATCH SIMPLE POR CÉDULA (para los NO conflictivos)
 #--------------------------------------------------------
 ids_aprobados = set(aprobados["NumeroID"]) - {""}
 match_id_simple = df["NumeroID"].isin(ids_aprobados) & (df["NumeroID"] != "")
 #--------------------------------------------------------
 # RESULTADO FINAL
 #--------------------------------------------------------
 aprobo = np.where(es_conflictivo, match_id_correo, match_id_simple)
 df["Nota"] = ""
 df["Aprobó"] = np.where(aprobo, "Sí", "No")
 #--------------------------------------------------------
 # MARCAR PARA REVISIÓN: cédula ambigua, aparece en Aprobados,
 # pero el correo no coincide con nadie -> no se puede confirmar
 #--------------------------------------------------------
 mascara_inconsistente = (
     es_conflictivo
& df["NumeroID"].isin(ids_aprobados)
& (~match_id_correo)
 )
 df.loc[mascara_inconsistente, "Aprobó"] = "Revisar"
 agregar_observacion(
     df,
     mascara_inconsistente,
     "Cédula duplicada en Participantes - revisar manualmente"
 )
 return df
# ==========================================================
# MOTOR GENERAL
# ==========================================================
def procesar_resultados(
  participantes,
  resultados,
  tipo_curso
):
  if tipo_curso == "Con nota":
      return procesar_con_nota(
          participantes,
          resultados
      )
  return procesar_sin_nota(
      participantes,
      resultados
  )
# ==========================================================
# CONSTRUIR REPORTE FINAL
# ==========================================================
def construir_reporte(df):
  reporte = df.copy()
  reporte = reporte.rename(
      columns={
          "Nombre": "Nombres y apellidos",
          "NumeroID": "Cédula"
      }
  )
  columnas = [
      "Nombres y apellidos",
      "Cédula",
      "Cargo",
      "Seccional",
      "Dependencia",
      "Nota",
      "Aprobó",
      "Observaciones"
  ]
  for columna in columnas:
      if columna not in reporte.columns:
          reporte[columna] = ""
  reporte = reporte[columnas]
  return reporte
# ==========================================================
# GENERAR RESUMEN
# ==========================================================
def generar_resumen(
  reporte,
  estadisticas_planta,
  estadisticas_validacion
):
  resumen = {}
  resumen["Total participantes"] = len(reporte)
  resumen["Total aprobados"] = int(
      (reporte["Aprobó"] == "Sí").sum()
  )
  resumen["Total reprobados"] = int(
      (reporte["Aprobó"] == "No").sum()
  )
  resumen["IDs vacíos"] = estadisticas_validacion[
      "IDs Vacíos"
  ]
  resumen["Duplicados"] = estadisticas_validacion[
      "Duplicados"
  ]
  resumen["Encontrados por ID"] = estadisticas_planta[
      "Encontrados por ID"
  ]
  resumen["Encontrados por Nombre"] = estadisticas_planta[
      "Encontrados por Nombre"
  ]
  resumen["Cédulas recuperadas"] = estadisticas_planta[
      "Cedulas Recuperadas"
  ]
  resumen["No encontrados en Planta"] = estadisticas_planta[
      "No encontrados en Planta"
  ]
  resumen["Nombres duplicados en Planta"] = estadisticas_planta[
      "Nombres Duplicados en Planta"
  ]
  resumen["Aprobados duplicados"] = contar_aprobados_duplicados(
      reporte.rename(columns={"Cédula": "NumeroID"})
  )
  resumen["Notas inconsistentes (revisar)"] = int(
      (reporte["Aprobó"] == "Revisar").sum()
  )
  return resumen
# ==========================================================
# EXPORTAR A EXCEL
# ==========================================================
def exportar_excel(reporte, resumen):
  wb = Workbook()
  # ======================================================
  # HOJA REPORTE
  # ======================================================
  ws = wb.active
  ws.title = "Reporte"
  for fila in dataframe_to_rows(
      reporte,
      index=False,
      header=True
  ):
      ws.append(fila)
  # ---------------------------------------------
  # Encabezados
  # ---------------------------------------------
  encabezado = PatternFill(
      fill_type="solid",
      fgColor=COLOR_VERDE
  )
  borde = Border(
      left=Side(style="thin"),
      right=Side(style="thin"),
      top=Side(style="thin"),
      bottom=Side(style="thin")
  )
  fuente = Font(
      bold=True
  )
  for celda in ws[1]:
      celda.fill = encabezado
      celda.font = fuente
      celda.border = borde
      celda.alignment = Alignment(
          horizontal="center"
      )
  # ---------------------------------------------
  # Pintar filas
  # ---------------------------------------------
  amarillo = PatternFill(
      fill_type="solid",
      fgColor=COLOR_AMARILLO
  )
  rojo = PatternFill(
      fill_type="solid",
      fgColor=COLOR_ROJO
  )
  col_obs = None
  for i, c in enumerate(ws[1], start=1):
      if c.value == "Observaciones":
          col_obs = i
          break
  if col_obs:
      for fila in ws.iter_rows(min_row=2):
          texto = str(
              fila[col_obs - 1].value
          )
          color = None
          if "Cédula duplicada" in texto:
              color = rojo
          elif "ID vacío" in texto:
              color = amarillo
          if color:
              for celda in fila:
                  celda.fill = color
  # ---------------------------------------------
  # Ajustar ancho columnas
  # ---------------------------------------------
  for columna in ws.columns:
      largo = max(
          len(str(c.value))
          if c.value is not None
          else 0
          for c in columna
      )
      ws.column_dimensions[
          columna[0].column_letter
      ].width = largo + 3
  # ======================================================
  # HOJA RESUMEN
  # ======================================================
  resumen_ws = wb.create_sheet("Resumen")
  resumen_ws.append([
      "Indicador",
      "Valor"
  ])
  resumen_ws["A1"].fill = encabezado
  resumen_ws["B1"].fill = encabezado
  resumen_ws["A1"].font = fuente
  resumen_ws["B1"].font = fuente
  for indicador, valor in resumen.items():
      resumen_ws.append([
          indicador,
          valor
      ])
  for columna in resumen_ws.columns:
      largo = max(
          len(str(c.value))
          if c.value is not None
          else 0
          for c in columna
      )
      resumen_ws.column_dimensions[
          columna[0].column_letter
      ].width = largo + 3
  # ======================================================
  # GUARDAR
  # ======================================================
  salida = BytesIO()
  wb.save(salida)
  salida.seek(0)
  return salida
# ==========================================================
# INTERFAZ
# ==========================================================
tipo_curso = st.selectbox(
  "Tipo de curso",
  [
      "Con nota",
      "Sin nota"
  ]
)
st.divider()
col1, col2 = st.columns(2)
with col1:
  participantes_file = st.file_uploader(
      "📄 Participantes",
      type="xlsx"
  )
  planta_file = st.file_uploader(
      "🏢 Planta",
      type="xlsx"
  )
with col2:
  if tipo_curso == "Con nota":
      resultados_file = st.file_uploader(
          "📝 Calificaciones",
          type="xlsx"
      )
  else:
      resultados_file = st.file_uploader(
          "✅ Aprobados",
          type="xlsx"
      )
st.divider()
# ==========================================================
# BOTÓN
# ==========================================================
if st.button(
  "🚀 Procesar",
  use_container_width=True
):
  if participantes_file is None:
      st.error("Debe cargar Participantes.")
      st.stop()
  if planta_file is None:
      st.error("Debe cargar Planta.")
      st.stop()
  if resultados_file is None:
      st.error("Debe cargar el archivo de resultados.")
      st.stop()
  barra = st.progress(0)
  # ------------------------------------------------------
  # Leer archivos
  # ------------------------------------------------------
  participantes = cargar_excel(
      participantes_file
  )
  planta = cargar_excel(
      planta_file
  )
  resultados = cargar_excel(
      resultados_file
  )
  barra.progress(10)
  # ------------------------------------------------------
  # Preparar
  # ------------------------------------------------------
  participantes = preparar_participantes(
      participantes
  )
  planta = preparar_planta(
      planta
  )
  if tipo_curso == "Con nota":
      resultados = preparar_calificaciones(
          resultados
      )
  else:
      resultados = preparar_aprobados(
          resultados
      )
  barra.progress(25)
  # ------------------------------------------------------
  # Cruce Planta
  # ------------------------------------------------------
  participantes, estadisticas_planta = enriquecer_con_planta(
      participantes,
      planta
  )
  barra.progress(45)
  # ------------------------------------------------------
  # Validaciones
  # ------------------------------------------------------
  participantes, estadisticas_validacion = ejecutar_validaciones(
      participantes
  )
  barra.progress(60)
  # ------------------------------------------------------
  # Resultados
  # ------------------------------------------------------
  participantes = procesar_resultados(
      participantes,
      resultados,
      tipo_curso
  )
  participantes = limpiar_observaciones(participantes)
  barra.progress(80)
  # ------------------------------------------------------
  # Reporte
  # ------------------------------------------------------
  reporte = construir_reporte(
      participantes
  )
  resumen = generar_resumen(
      reporte,
      estadisticas_planta,
      estadisticas_validacion
  )
  barra.progress(100)
  # ------------------------------------------------------
  # Guardar Session State
  # ------------------------------------------------------
  st.session_state.reporte = reporte
  st.session_state.resumen = resumen
  st.session_state.archivo = exportar_excel(
      reporte,
      resumen
  )
  st.success("Proceso finalizado correctamente.")
# ==========================================================
# RESULTADOS
# ==========================================================
if st.session_state.reporte is not None:
  st.divider()
  resumen = st.session_state.resumen
  c1, c2, c3, c4 = st.columns(4)
  c1.metric(
      "Participantes",
      resumen["Total participantes"]
  )
  c2.metric(
      "Aprobados",
      resumen["Total aprobados"]
  )
  c3.metric(
      "IDs Vacíos",
      resumen["IDs vacíos"]
  )
  c4.metric(
      "Duplicados",
      resumen["Duplicados"]
  )
  c1, c2, c3, c4 = st.columns(4)
  c1.metric(
      "Encontrados ID",
      resumen["Encontrados por ID"]
  )
  c2.metric(
      "Encontrados Nombre",
      resumen["Encontrados por Nombre"]
  )
  c3.metric(
      "Cédulas recuperadas",
      resumen["Cédulas recuperadas"]
  )
  c4.metric(
      "No encontrados",
      resumen["No encontrados en Planta"]
  )
  st.divider()
  st.subheader("Vista previa")
  st.dataframe(
      st.session_state.reporte,
      use_container_width=True,
      height=500
  )
  st.download_button(
      "📥 Descargar Excel",
      data=st.session_state.archivo,
      file_name="Reporte_Final.xlsx",
      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
      use_container_width=True
  )